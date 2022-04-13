import time
import datetime
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from auth_data import id_password

path = "123.xlsx"  # имя файла

n = input(str("Введите номер протокола: "))
date_prot = input('введите дату проведения протокола (гггг-мм-дд): ')
path_protokol = input('введите путь к протоколу: ')

date_now = datetime.date.today()
date_prot = date_prot.split('-')
day_count = datetime.date(int(date_prot[0]), int(date_prot[1]), int(date_prot[2]))
day_count = date_now - day_count
day_count = str(day_count)
day_count = (day_count.split()[0])


wb_obj = openpyxl.load_workbook(path)  # Открываем файл
sheet_obj = wb_obj.active  # Выбираем активный лист таблицы(
m_row = sheet_obj.max_row

s = Service("/home/odinokov/PycharmProjects/pythonProject2/chromedriver")

url = "https://ideas.nlmk.com/"

options = webdriver.ChromeOptions()
options.add_argument("user-agent=Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:84.0) Gecko/20100101 Firefox/84.0")
options.headless = False

driver = webdriver.Chrome(
    service=s,
    options=options
)

def iogin():
    email_input = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "login")))
    email_input.clear()
    email_input.send_keys("odinokov_da")
    password_input = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "password")))
    password_input.clear()
    password_input.send_keys(id_password)
    password_input.send_keys(Keys.ENTER)

def find_elment_class(n):
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, n))).click()

def find_element_link(n):
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, n))).click()

def find_element_xpaths(n):
    ds = driver.find_elements(By.XPATH, n)
    return ds

def find_element_xpath(n):
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, n))).click()

def find_element_xpath_DOWN(n):
    driver.find_element(By.XPATH,n).send_keys(Keys.DOWN)
    time.sleep(3)

def find_element_xpath_prot(n, path_protokol):
    driver.find_element(By.XPATH, n).send_keys(path_protokol)

def find_id(n):
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, n))).click()

def find_iframe(n):
    driver.switch_to.frame(driver.find_element(By.TAG_NAME, n))
    time.sleep(4)

def send_prot(s,n):
    input_n_prot = driver.find_element(By.XPATH, s)
    input_n_prot.clear()
    input_n_prot.send_keys(n)
    time.sleep(3)

def send_date(n, day_count):
    driver.find_element(By.XPATH, n).clear()
    driver.find_element(By.XPATH, n).send_keys(Keys.LEFT * int(day_count))
    time.sleep(2)

def send_date2(n, day_count):
    driver.find_element(By.XPATH, n).clear()
    driver.find_element(By.XPATH, n).send_keys(Keys.LEFT * int(day_count))
    time.sleep(2)

try:
    driver.get(url=url)
    time.sleep(15)
    iogin()
    print('вход')
    find_elment_class('work')
    find_element_link("Все идеи")
    print('все')
    find_element_link("2ТС без ТЭ")
    print('2ТС ТЭ')

    my_ul2 = driver.find_elements(By.XPATH, "//ul[@class='b-page']")  # блок со страницам
    if len(my_ul2) == 1:
        my_ul = driver.find_element(By.XPATH, "//ul[@class='b-page']")  # блок со страницам
        time.sleep(7)
        all_li = my_ul.find_elements(By.TAG_NAME, "li")  # оперделяем кол-во страниц
        time.sleep(7)
        link_list = []
        for li in all_li:
            y = li.text
            link_list.append(y)
        rec = 1
        while rec == 1:
            for i in range(2, m_row + 1):
                cell_obj = sheet_obj.cell(row=i, column=1)  # В column= подставляем номер нужной колонки
                number_ideas = cell_obj.value
                cell_ob = sheet_obj.cell(row=i, column=10)  # В column= подставляем номер нужной колонки
                name_ideas = cell_ob.value
                for y in link_list:
                    ds = driver.find_elements(By.XPATH, " // *[ @ href = 'Offer.aspx?id=" + str(number_ideas) + "']")
                    time.sleep(5)

                    if len(ds) >= 1:
                        find_element_xpath(" // *[ @ href = 'Offer.aspx?id=" + str(number_ideas) + "']")
                        find_id("ctl00_ctl00_ctl00_MainContent_VMenuRightContent_OfferMenu_linkEngine83")
                        find_id("ctl00_ctl00_ctl00_MainContent_CommandsMenu_linkEdit")
                        driver.switch_to.frame(driver.find_element(By.TAG_NAME, "iframe"))
                        time.sleep(5)
                        send_prot("// input[@class='alpaca-control form-control' and @id = 'alpaca2']", n)
                        send_date("// input[@class='alpaca-control form-control' and @id = 'alpaca3']", day_count)
                        send_date2("// input[@class='alpaca-control form-control' and @id = 'alpaca4']", day_count)
                        driver.switch_to.default_content()
                        time.sleep(3)
                        find_id("ctl00_ctl00_ctl00_MainContent_CommandsMenu_linkSave")
                        find_id("ctl00_ctl00_ctl00_MainContent_VMenuRightContent_"
                                                "OfferMenu_linkDocuments")
                        find_element_xpath_prot("//input[@type='file'][contains(@id,'FileUpload1')]", path_protokol)
                        find_element_xpath("//input[@value='Прикрепить файл']")
                        """
                        find_id("ctl00_ctl00_ctl00_MainContent_VMenuRightContent_OfferMenu_linkStatus")
                        find_id("ctl00_ctl00_ctl00_MainContent_CommandsMenu_linkEdit")
                        find_element_xpath_DOWN("(//select[@class='ddlStatusList'])[2]")
                        find_element_xpath("//input[@value='Изменить статус']")
                        """
                        find_element_link("Все идеи")
                        print(name_ideas)
                        break
                    else:
                        rec = 0
                        w = int(y) + 1
                        print(f"переходим на страницу: {w}")
                        driver.find_element(By.LINK_TEXT,""+str(w)+"").click()
                        time.sleep(3)

    else:
        for i in range(2, m_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=1)  # В column= подставляем номер нужной колонки
            number_ideas = cell_obj.value
            cell_ob = sheet_obj.cell(row=i, column=10)  # В column= подставляем номер нужной колонки
            name_ideas = cell_ob.value
            find_element_xpath(" // *[ @ href = 'Offer.aspx?id=" + str(number_ideas) + "']")
            find_id("ctl00_ctl00_ctl00_MainContent_VMenuRightContent_OfferMenu_linkEngine83")
            find_id("ctl00_ctl00_ctl00_MainContent_CommandsMenu_linkEdit")
            driver.switch_to.frame(driver.find_element(By.TAG_NAME, "iframe"))
            time.sleep(5)
            send_prot("// input[@class='alpaca-control form-control' and @id = 'alpaca2']", n)
            send_date("// input[@class='alpaca-control form-control' and @id = 'alpaca3']", day_count)
            send_date2("// input[@class='alpaca-control form-control' and @id = 'alpaca4']", day_count)
            driver.switch_to.default_content()
            time.sleep(3)
            find_id("ctl00_ctl00_ctl00_MainContent_CommandsMenu_linkSave")
            find_id("ctl00_ctl00_ctl00_MainContent_VMenuRightContent_"
                                    "OfferMenu_linkDocuments")
            find_element_xpath_prot("//input[@type='file'][contains(@id,'FileUpload1')]", path_protokol)
            find_element_xpath("//input[@value='Прикрепить файл']")
            """
            find_id("ctl00_ctl00_ctl00_MainContent_VMenuRightContent_OfferMenu_linkStatus")
            find_id("ctl00_ctl00_ctl00_MainContent_CommandsMenu_linkEdit")
            find_element_xpath_DOWN("(//select[@class='ddlStatusList'])[2]")
            find_element_xpath("//input[@value='Изменить статус']")
            """
            find_element_link("Все идеи")
            print(name_ideas)

except Exception as ex:
    print(ex)
finally:
    driver.close()
    driver.quit()
